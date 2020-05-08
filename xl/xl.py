import openpyxl


# return a list
# input excel filename, first_item(eg:A1), end_item(eg:A10), sheet_name,return a list
def xl_list(filename, first_item, end_item, sheet_name):
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheet_name]
    list_cell = []
    for row in ws[first_item:end_item]:
        for cell in row:
            list_cell.append(cell.value)

    return list_cell


# return a EXCEL file
# input filename(excel name,eg:1.xlsx), list_name, list_num(eg:A eg:B)
def list_xl(filename, list_name, list_num):
    wb = openpyxl.Workbook()
    ws = wb[wb.sheetnames[0]]

    list_len = len(list_name)
    sheet_num = 1

    while list_len >= sheet_num:
        sheet_name = '{}{}'.format(list_num, sheet_num)
        ws[sheet_name] = list_name[sheet_num - 1]
        sheet_num += 1

    wb.save(filename)


# return a dir
# input filename(excel name,eg:1.xlsx), list_one(eg:A), list_two(eg:B)
def xl_dir(filename, list_one, list_two):
    wb = openpyxl.load_workbook(filename)
    ws = wb[wb.sheetnames[0]]

    max_row = ws.max_row
    row = 1
    dir_name = {}
    while row <= max_row:
        new_list_one = list_one
        new_list_one = new_list_one + '{}'.format(row)

        new_list_two = list_two
        new_list_two = new_list_two + '{}'.format(row)
        dir_name[ws[new_list_one].value] = ws[new_list_two].value
        row += 1

    return dir_name



